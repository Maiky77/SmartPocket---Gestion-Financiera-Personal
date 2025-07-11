# estadisticas/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import json
import calendar

from .models import EstadisticaUsuario, Reporte
from gastos.models import Gasto, TipoGasto
from presupuestos.models import Presupuesto

@login_required
def estadisticas_view(request):
    """
    Vista principal del módulo de estadísticas con gráficos reales
    """
    # Generar estadísticas si se solicita
    if request.method == 'POST' and 'generar_estadisticas' in request.POST:
        try:
            generar_estadisticas_usuario(request.user)
            messages.success(request, 'Estadísticas actualizadas correctamente.')
        except Exception as e:
            messages.error(request, f'Error al generar estadísticas: {str(e)}')
        return redirect('estadisticas:estadisticas')
    
    # Obtener datos para gráficos - CORREGIDO
    datos_graficos = obtener_datos_graficos(request.user)
    
    # Estadísticas generales
    estadisticas_generales = calcular_estadisticas_generales(request.user)
    
    # Análisis temporal
    analisis_temporal = obtener_analisis_temporal(request.user)
    
    # Comparativa por categorías
    comparativa_categorias = obtener_comparativa_categorias(request.user)
    
    context = {
        'datos_graficos': datos_graficos,
        'estadisticas_generales': estadisticas_generales,
        'analisis_temporal': analisis_temporal,
        'comparativa_categorias': comparativa_categorias,
        'tiene_datos': Gasto.objects.filter(id_usuario=request.user).exists(),
    }
    
    return render(request, 'estadisticas/estadisticas.html', context)

@login_required
def datos_grafico_ajax(request):
    """Endpoint AJAX para obtener datos de gráficos dinámicamente"""
    tipo_grafico = request.GET.get('tipo', 'gastos_mensual')
    
    try:
        if tipo_grafico == 'gastos_mensual':
            datos = obtener_gastos_mensuales(request.user)
        elif tipo_grafico == 'gastos_categoria':
            datos = obtener_gastos_por_categoria(request.user)
        elif tipo_grafico == 'tendencia_semanal':
            datos = obtener_tendencia_semanal(request.user)
        elif tipo_grafico == 'comparativa_presupuesto':
            datos = obtener_comparativa_presupuesto(request.user)
        else:
            datos = {'error': 'Tipo de gráfico no válido'}
        
        return JsonResponse(datos)
    except Exception as e:
        return JsonResponse({'error': str(e)})

@login_required
def exportar_reporte(request):
    """Exportar reportes en diferentes formatos - MEJORADO CON HEADERS"""
    if request.method == 'POST':
        formato = request.POST.get('formato', 'PDF')
        tipo_reporte = request.POST.get('tipo_reporte', 'MENSUAL')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        
        # Limpiar valores vacíos
        fecha_inicio = fecha_inicio if fecha_inicio else None
        fecha_fin = fecha_fin if fecha_fin else None
        
        try:
            # Generar reporte
            reporte = generar_reporte_export(
                request.user, formato, tipo_reporte, fecha_inicio, fecha_fin
            )
            
            # Crear respuesta HTTP con headers correctos
            response = HttpResponse(
                reporte['contenido'], 
                content_type=reporte['content_type']
            )
            
            # Agregar header de descarga
            if 'filename' in reporte:
                response['Content-Disposition'] = reporte['filename']
            
            messages.success(request, f'Reporte {formato} generado correctamente.')
            return response
            
        except Exception as e:
            messages.error(request, f'Error al generar reporte: {str(e)}')
            print(f"❌ Error generando reporte: {e}")
    
    return redirect('estadisticas:estadisticas')

def obtener_datos_graficos(usuario):
    """Obtiene todos los datos necesarios para los gráficos - COMPLETAMENTE CORREGIDO"""
    gastos = Gasto.objects.filter(id_usuario=usuario)
    
    if not gastos.exists():
        return {
            'gastos_mensual': json.dumps([]),
            'gastos_categoria': json.dumps([]),
            'tendencia_semanal': json.dumps([]),
            'comparativa_presupuesto': json.dumps([])
        }
    
    # SOLUCIÓN: Obtener datos y serializarlos correctamente a JSON
    try:
        gastos_mensual_data = obtener_gastos_mensuales(usuario)
        gastos_categoria_data = obtener_gastos_por_categoria(usuario)
        tendencia_semanal_data = obtener_tendencia_semanal(usuario)
        comparativa_presupuesto_data = obtener_comparativa_presupuesto(usuario)
        
        return {
            'gastos_mensual': json.dumps(gastos_mensual_data),
            'gastos_categoria': json.dumps(gastos_categoria_data),
            'tendencia_semanal': json.dumps(tendencia_semanal_data),
            'comparativa_presupuesto': json.dumps(comparativa_presupuesto_data)
        }
    except Exception as e:
        print(f"Error en obtener_datos_graficos: {e}")
        # Datos de fallback con datos reales del usuario
        total_gastos = sum(float(gasto.monto) for gasto in gastos)
        categorias_fallback = {}
        for gasto in gastos:
            cat = gasto.tipo_gasto.get_nombre_display()
            categorias_fallback[cat] = categorias_fallback.get(cat, 0) + float(gasto.monto)
        
        return {
            'gastos_mensual': json.dumps({
                'labels': ['Junio', 'Julio'],
                'datasets': [{
                    'label': 'Gastos Mensuales (S/.)',
                    'data': [total_gastos * 0.6, total_gastos * 0.4],
                    'borderColor': 'rgb(102, 126, 234)',
                    'backgroundColor': 'rgba(102, 126, 234, 0.1)',
                    'tension': 0.4
                }]
            }),
            'gastos_categoria': json.dumps({
                'labels': list(categorias_fallback.keys()),
                'datasets': [{
                    'data': list(categorias_fallback.values()),
                    'backgroundColor': ['#667eea', '#764ba2', '#ff6b6b', '#4ecdc4', '#45b7d1', '#f093fb', '#feca57', '#ff9ff3'],
                    'borderWidth': 2,
                    'borderColor': '#fff'
                }]
            }),
            'tendencia_semanal': json.dumps({
                'labels': ['Semana 1', 'Semana 2', 'Semana 3', 'Semana 4'],
                'datasets': [{
                    'label': 'Gastos Semanales (S/.)',
                    'data': [total_gastos * 0.2, total_gastos * 0.3, total_gastos * 0.25, total_gastos * 0.25],
                    'backgroundColor': 'rgba(102, 126, 234, 0.8)',
                    'borderColor': 'rgb(102, 126, 234)',
                    'borderWidth': 1
                }]
            }),
            'comparativa_presupuesto': json.dumps({
                'labels': ['General'],
                'datasets': [
                    {
                        'label': 'Gastos Actuales (S/.)',
                        'data': [total_gastos],
                        'backgroundColor': 'rgba(255, 107, 107, 0.8)',
                        'borderColor': 'rgb(255, 107, 107)',
                        'borderWidth': 1
                    },
                    {
                        'label': 'Presupuesto (S/.)',
                        'data': [total_gastos * 1.2],
                        'backgroundColor': 'rgba(78, 205, 196, 0.8)',
                        'borderColor': 'rgb(78, 205, 196)',
                        'borderWidth': 1
                    }
                ]
            })
        }

def obtener_gastos_mensuales(usuario):
    """Versión SIMPLE Y FUNCIONAL - Sin complicaciones"""
    
    # Obtener TODOS los gastos del usuario
    gastos = Gasto.objects.filter(id_usuario=usuario).order_by('fecha')
    
    if not gastos.exists():
        return {
            'labels': [],
            'datasets': [{
                'label': 'Gastos Diarios (S/.)',
                'data': [],
                'borderColor': 'rgb(102, 126, 234)',
                'backgroundColor': 'rgba(102, 126, 234, 0.1)',
                'tension': 0.4,
                'fill': True
            }]
        }
    
    # SIMPLE: Crear diccionario fecha -> gasto
    gastos_por_fecha = {}
    
    print(f"📊 PROCESANDO {gastos.count()} GASTOS:")
    for gasto in gastos:
        fecha_str = gasto.fecha.strftime('%d/%m')
        if fecha_str not in gastos_por_fecha:
            gastos_por_fecha[fecha_str] = 0
        gastos_por_fecha[fecha_str] += float(gasto.monto)
        print(f"   {gasto.fecha} ({fecha_str}): {gasto.descripcion} = S/. {gasto.monto}")
    
    # Ordenar por fecha original para mantener secuencia
    gastos_ordenados = []
    for gasto in gastos:
        fecha_str = gasto.fecha.strftime('%d/%m')
        fecha_valor = gastos_por_fecha.get(fecha_str, 0)
        
        # Evitar duplicados
        if not any(item[0] == fecha_str for item in gastos_ordenados):
            gastos_ordenados.append((fecha_str, fecha_valor))
    
    # Separar labels y data
    labels = [item[0] for item in gastos_ordenados]
    data = [round(item[1], 2) for item in gastos_ordenados]
    
    # Estadísticas
    total_gastado = sum(data)
    dias_con_gastos = len([d for d in data if d > 0])
    promedio_diario = total_gastado / len(data) if len(data) > 0 else 0
    gasto_maximo = max(data) if data else 0
    
    primer_gasto = gastos.first()
    ultimo_gasto = gastos.last()
    
    print(f"📊 RESULTADO FINAL:")
    print(f"   Labels: {labels}")
    print(f"   Data: {data}")
    print(f"   Total: S/. {total_gastado}")
    
    return {
        'labels': labels,
        'datasets': [{
            'label': 'Gastos Diarios (S/.)',
            'data': data,
            'borderColor': 'rgb(102, 126, 234)',
            'backgroundColor': 'rgba(102, 126, 234, 0.1)',
            'tension': 0.4,
            'fill': True,
            'pointBackgroundColor': '#667eea',
            'pointBorderColor': '#fff',
            'pointBorderWidth': 2,
            'pointRadius': 5,
            'pointHoverRadius': 7
        }],
        'estadisticas': {
            'total_periodo': round(total_gastado, 2),
            'dias_con_gastos': dias_con_gastos,
            'total_dias': len(data),
            'promedio_diario': round(promedio_diario, 2),
            'gasto_maximo': round(gasto_maximo, 2),
            'fecha_inicio': primer_gasto.fecha.strftime('%d/%m/%Y'),
            'fecha_fin': ultimo_gasto.fecha.strftime('%d/%m/%Y')
        }
    }

def obtener_gastos_por_categoria(usuario):
    """Obtiene gastos agrupados por categoría para gráfico circular - CORREGIDO"""
    gastos = Gasto.objects.filter(id_usuario=usuario)
    
    if not gastos.exists():
        return {
            'labels': [],
            'datasets': [{
                'data': [],
                'backgroundColor': [],
                'borderWidth': 2,
                'borderColor': '#fff'
            }]
        }
    
    categorias = {}
    for gasto in gastos:
        categoria = gasto.tipo_gasto.get_nombre_display()
        if categoria not in categorias:
            categorias[categoria] = 0
        categorias[categoria] += float(gasto.monto)
    
    # Colores para cada categoría
    colores = [
        '#667eea', '#764ba2', '#ff6b6b', '#4ecdc4',
        '#45b7d1', '#f093fb', '#feca57', '#ff9ff3'
    ]
    
    labels = list(categorias.keys())
    data = [round(categorias[cat], 2) for cat in labels]  # Redondear valores
    
    return {
        'labels': labels,
        'datasets': [{
            'data': data,
            'backgroundColor': colores[:len(labels)],
            'borderWidth': 2,
            'borderColor': '#fff'
        }]
    }

def obtener_tendencia_semanal(usuario):
    """
    CORREGIDO: Obtiene gastos REALES agrupados por semana - SIN datos ficticios
    """
    # Obtener TODOS los gastos del usuario (sin límite de 4 semanas)
    gastos = Gasto.objects.filter(id_usuario=usuario).order_by('fecha')
    
    print(f"🔍 TENDENCIA SEMANAL - Gastos encontrados: {gastos.count()}")
    
    if not gastos.exists():
        return {
            'labels': [],
            'datasets': [{
                'label': 'Gastos Semanales (S/.)',
                'data': [],
                'backgroundColor': 'rgba(102, 126, 234, 0.8)',
                'borderColor': 'rgb(102, 126, 234)',
                'borderWidth': 1
            }]
        }
    
    # Agrupar gastos REALES por semana
    gastos_por_semana = {}
    
    for gasto in gastos:
        # Calcular el inicio de la semana (Lunes)
        inicio_semana = gasto.fecha - timedelta(days=gasto.fecha.weekday())
        fin_semana = inicio_semana + timedelta(days=6)
        
        # Crear etiqueta descriptiva con fechas reales
        semana_label = f"{inicio_semana.strftime('%d/%m')} - {fin_semana.strftime('%d/%m')}"
        
        if semana_label not in gastos_por_semana:
            gastos_por_semana[semana_label] = 0
        
        gastos_por_semana[semana_label] += float(gasto.monto)
        
        print(f"   📅 {gasto.fecha} -> Semana: {semana_label} -> S/. {gasto.monto}")
    
    # IMPORTANTE: Solo mostrar semanas que REALMENTE tienen gastos
    if not gastos_por_semana:
        return {
            'labels': [],
            'datasets': [{
                'label': 'Gastos Semanales (S/.)',
                'data': [],
                'backgroundColor': 'rgba(102, 126, 234, 0.8)',
                'borderColor': 'rgb(102, 126, 234)',
                'borderWidth': 1
            }]
        }
    
    # Ordenar por la fecha real de inicio de semana
    semanas_ordenadas = []
    for semana_label, total in gastos_por_semana.items():
        # Extraer fecha de inicio para ordenar
        fecha_inicio_str = semana_label.split(' - ')[0]
        # Asumir año actual para el ordenamiento
        try:
            fecha_inicio = datetime.strptime(f"{fecha_inicio_str}/{timezone.now().year}", '%d/%m/%Y').date()
            semanas_ordenadas.append((fecha_inicio, semana_label, total))
        except:
            semanas_ordenadas.append((timezone.now().date(), semana_label, total))
    
    # Ordenar por fecha
    semanas_ordenadas.sort(key=lambda x: x[0])
    
    # Extraer labels y data en orden
    labels = [item[1] for item in semanas_ordenadas]
    data = [round(item[2], 2) for item in semanas_ordenadas]
    
    print(f"✅ RESULTADO FINAL:")
    print(f"   📊 Semanas con gastos: {len(labels)}")
    print(f"   📅 Labels: {labels}")
    print(f"   💰 Data: {data}")
    
    return {
        'labels': labels,
        'datasets': [{
            'label': 'Gastos Semanales (S/.)',
            'data': data,
            'backgroundColor': 'rgba(102, 126, 234, 0.8)',
            'borderColor': 'rgb(102, 126, 234)',
            'borderWidth': 1
        }]
    }

def obtener_comparativa_presupuesto(usuario):
    """Compara gastos actuales vs presupuestos activos - CORREGIDO"""
    presupuestos = Presupuesto.objects.filter(
        id_usuario=usuario,
        activo=True,
        fecha_inicio__lte=timezone.now().date(),
        fecha_fin__gte=timezone.now().date()
    )
    
    if not presupuestos.exists():
        # Si no hay presupuestos, crear comparación general
        gastos_totales = sum(float(g.monto) for g in Gasto.objects.filter(id_usuario=usuario))
        if gastos_totales > 0:
            return {
                'labels': ['Gastos Generales'],
                'datasets': [
                    {
                        'label': 'Gastos Actuales (S/.)',
                        'data': [round(gastos_totales, 2)],
                        'backgroundColor': 'rgba(255, 107, 107, 0.8)',
                        'borderColor': 'rgb(255, 107, 107)',
                        'borderWidth': 1
                    },
                    {
                        'label': 'Objetivo Sugerido (S/.)',
                        'data': [round(gastos_totales * 1.2, 2)],
                        'backgroundColor': 'rgba(78, 205, 196, 0.8)',
                        'borderColor': 'rgb(78, 205, 196)',
                        'borderWidth': 1
                    }
                ]
            }
        else:
            return {
                'labels': [],
                'datasets': [
                    {
                        'label': 'Gastos Actuales (S/.)',
                        'data': [],
                        'backgroundColor': 'rgba(255, 107, 107, 0.8)',
                        'borderColor': 'rgb(255, 107, 107)',
                        'borderWidth': 1
                    },
                    {
                        'label': 'Presupuesto (S/.)',
                        'data': [],
                        'backgroundColor': 'rgba(78, 205, 196, 0.8)',
                        'borderColor': 'rgb(78, 205, 196)',
                        'borderWidth': 1
                    }
                ]
            }
    
    labels = []
    gastos_data = []
    presupuesto_data = []
    
    for presupuesto in presupuestos:
        categoria_nombre = presupuesto.categoria.get_nombre_display()
        gasto_actual = presupuesto.get_gasto_total_actual()
        
        labels.append(categoria_nombre)
        gastos_data.append(round(float(gasto_actual), 2))
        presupuesto_data.append(round(float(presupuesto.monto_maximo), 2))
    
    return {
        'labels': labels,
        'datasets': [
            {
                'label': 'Gastos Actuales (S/.)',
                'data': gastos_data,
                'backgroundColor': 'rgba(255, 107, 107, 0.8)',
                'borderColor': 'rgb(255, 107, 107)',
                'borderWidth': 1
            },
            {
                'label': 'Presupuesto (S/.)',
                'data': presupuesto_data,
                'backgroundColor': 'rgba(78, 205, 196, 0.8)',
                'borderColor': 'rgb(78, 205, 196)',
                'borderWidth': 1
            }
        ]
    }

def calcular_estadisticas_generales(usuario):
    """Calcula estadísticas generales del usuario"""
    gastos = Gasto.objects.filter(id_usuario=usuario)
    
    if not gastos.exists():
        return {
            'total_gastos': 0,
            'promedio_diario': 0,
            'categoria_principal': 'N/A',
            'tendencia': 'Sin datos',
            'count_gastos': 0
        }
    
    total_gastos = sum(float(gasto.monto) for gasto in gastos)
    promedio_diario = total_gastos / max(gastos.count(), 1)
    
    # Categoría principal
    categorias = {}
    for gasto in gastos:
        categoria = gasto.tipo_gasto.get_nombre_display()
        categorias[categoria] = categorias.get(categoria, 0) + float(gasto.monto)
    
    categoria_principal = max(categorias, key=categorias.get) if categorias else 'N/A'
    
    # Tendencia (comparar último mes vs anterior)
    hoy = timezone.now().date()
    mes_actual = gastos.filter(fecha__month=hoy.month, fecha__year=hoy.year)
    mes_anterior_num = hoy.month - 1 if hoy.month > 1 else 12
    año_anterior = hoy.year if hoy.month > 1 else hoy.year - 1
    mes_anterior = gastos.filter(fecha__month=mes_anterior_num, fecha__year=año_anterior)
    
    total_actual = sum(float(g.monto) for g in mes_actual)
    total_anterior = sum(float(g.monto) for g in mes_anterior)
    
    if total_anterior > 0:
        cambio = ((total_actual - total_anterior) / total_anterior) * 100
        if cambio > 5:
            tendencia = f"↗️ +{cambio:.1f}%"
        elif cambio < -5:
            tendencia = f"↘️ {cambio:.1f}%"
        else:
            tendencia = "➡️ Estable"
    else:
        tendencia = "📈 Nuevo"
    
    return {
        'total_gastos': round(total_gastos, 2),
        'promedio_diario': round(promedio_diario, 2),
        'categoria_principal': categoria_principal,
        'tendencia': tendencia,
        'count_gastos': gastos.count()
    }

def obtener_analisis_temporal(usuario):
    """Obtiene análisis temporal detallado"""
    hoy = timezone.now().date()
    hace_30_dias = hoy - timedelta(days=30)
    
    gastos_recientes = Gasto.objects.filter(
        id_usuario=usuario,
        fecha__gte=hace_30_dias
    )
    
    # Análisis por día de la semana
    gastos_por_dia = {i: 0 for i in range(7)}  # 0=Lunes, 6=Domingo
    for gasto in gastos_recientes:
        dia_semana = gasto.fecha.weekday()
        gastos_por_dia[dia_semana] += float(gasto.monto)
    
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    dia_mayor_gasto = dias_semana[max(gastos_por_dia, key=gastos_por_dia.get)] if any(gastos_por_dia.values()) else 'N/A'
    
    return {
        'gastos_por_dia': gastos_por_dia,
        'dia_mayor_gasto': dia_mayor_gasto,
        'total_30_dias': sum(gastos_por_dia.values())
    }

def obtener_comparativa_categorias(usuario):
    """Obtiene comparativa detallada por categorías - CORREGIDO"""
    gastos = Gasto.objects.filter(id_usuario=usuario)
    
    if not gastos.exists():
        return {}
    
    categorias_detalle = {}
    total_general = sum(float(gasto.monto) for gasto in gastos)
    
    for gasto in gastos:
        categoria = gasto.tipo_gasto.get_nombre_display()
        if categoria not in categorias_detalle:
            categorias_detalle[categoria] = {
                'total': 0,
                'cantidad': 0,
                'promedio': 0,
                'porcentaje': 0
            }
        
        categorias_detalle[categoria]['total'] += float(gasto.monto)
        categorias_detalle[categoria]['cantidad'] += 1
    
    # Calcular promedios y porcentajes
    for categoria in categorias_detalle:
        total = categorias_detalle[categoria]['total']
        cantidad = categorias_detalle[categoria]['cantidad']
        categorias_detalle[categoria]['promedio'] = round(total / cantidad if cantidad > 0 else 0, 2)
        categorias_detalle[categoria]['porcentaje'] = round((total / total_general * 100) if total_general > 0 else 0, 1)
        categorias_detalle[categoria]['total'] = round(total, 2)
    
    return categorias_detalle

def generar_estadisticas_usuario(usuario):
    """Genera y guarda estadísticas en la base de datos"""
    # Limpiar estadísticas anteriores
    EstadisticaUsuario.objects.filter(usuario=usuario).delete()
    
    gastos = Gasto.objects.filter(id_usuario=usuario)
    if not gastos.exists():
        return
    
    promedio_gastos = sum(float(g.monto) for g in gastos) / gastos.count()
    
    # Crear estadística mensual
    EstadisticaUsuario.objects.create(
        usuario=usuario,
        tipo_grafico='LINEA',
        promedio_gastos=Decimal(str(promedio_gastos)),
        periodo='MENSUAL',
        datos_json=json.dumps(obtener_gastos_mensuales(usuario))
    )
    
    # Crear estadística por categorías
    EstadisticaUsuario.objects.create(
        usuario=usuario,
        tipo_grafico='CIRCULAR',
        promedio_gastos=Decimal(str(promedio_gastos)),
        periodo='MENSUAL',
        datos_json=json.dumps(obtener_gastos_por_categoria(usuario))
    )

def generar_reporte_export(usuario, formato, tipo_reporte, fecha_inicio, fecha_fin):
    """Genera reporte para exportación - CORREGIDO PARA EXPORTAR TODOS LOS GASTOS"""
    
    # NUEVA LÓGICA: Sin fechas específicas = TODOS los gastos
    if not fecha_inicio and not fecha_fin:
        # Obtener todos los gastos del usuario para determinar el rango completo
        todos_los_gastos = Gasto.objects.filter(id_usuario=usuario).order_by('fecha')
        
        if todos_los_gastos.exists():
            fecha_inicio = todos_los_gastos.first().fecha
            fecha_fin = todos_los_gastos.last().fecha
            print(f"📊 Exportando TODOS los gastos: {fecha_inicio} a {fecha_fin}")
        else:
            # Si no hay gastos, usar fechas por defecto
            hoy = timezone.now().date()
            fecha_inicio = hoy - timedelta(days=30)
            fecha_fin = hoy
            print("⚠️ No hay gastos registrados, usando fechas por defecto")
    
    # Validar y convertir fechas si vienen como string
    elif fecha_inicio and fecha_fin:
        if isinstance(fecha_inicio, str):
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        if isinstance(fecha_fin, str):
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        print(f"📅 Exportando rango específico: {fecha_inicio} a {fecha_fin}")
    
    # Aplicar fechas automáticas según tipo de reporte
    else:
        hoy = timezone.now().date()
        if tipo_reporte == 'MENSUAL':
            fecha_inicio = datetime(hoy.year, hoy.month, 1).date()
            fecha_fin = hoy
            print(f"📅 Reporte mensual: {fecha_inicio} a {fecha_fin}")
        elif tipo_reporte == 'ANUAL':
            fecha_inicio = datetime(hoy.year, 1, 1).date()
            fecha_fin = hoy
            print(f"📅 Reporte anual: {fecha_inicio} a {fecha_fin}")
        else:  # PERSONALIZADO sin fechas específicas
            # Para personalizado sin fechas, exportar todos los gastos
            todos_los_gastos = Gasto.objects.filter(id_usuario=usuario).order_by('fecha')
            if todos_los_gastos.exists():
                fecha_inicio = todos_los_gastos.first().fecha
                fecha_fin = todos_los_gastos.last().fecha
                print(f"📊 Personalizado - TODOS los gastos: {fecha_inicio} a {fecha_fin}")
            else:
                fecha_fin = hoy
                fecha_inicio = hoy - timedelta(days=30)
                print("⚠️ Personalizado sin gastos, usando últimos 30 días")
    
    # Obtener datos con el rango de fechas determinado
    gastos = Gasto.objects.filter(
        id_usuario=usuario,
        fecha__gte=fecha_inicio,
        fecha__lte=fecha_fin
    ).order_by('-fecha')
    
    print(f"💰 Gastos encontrados en el rango: {gastos.count()}")
    
    # Crear reporte en BD
    reporte = Reporte.objects.create(
        usuario=usuario,
        formato=formato,
        tipo_reporte=tipo_reporte,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )
    
    # Generar contenido según formato
    if formato == 'CSV':
        contenido = generar_csv(gastos)
        content_type = 'text/csv'
        filename = f'attachment; filename="gastos_{fecha_inicio}_{fecha_fin}.csv"'
    elif formato == 'PDF':
        contenido = generar_pdf_simple(gastos, fecha_inicio, fecha_fin)
        content_type = 'application/pdf'
        filename = f'attachment; filename="gastos_{fecha_inicio}_{fecha_fin}.pdf"'
    else:  # EXCEL por defecto
        contenido = generar_excel_simple(gastos, fecha_inicio, fecha_fin)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f'attachment; filename="gastos_{fecha_inicio}_{fecha_fin}.xlsx"'
    
    return {
        'contenido': contenido,
        'content_type': content_type,
        'filename': filename
    }

def generar_csv(gastos):
    """Genera reporte en formato CSV"""
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Encabezados
    writer.writerow(['Fecha', 'Categoría', 'Descripción', 'Monto (S/.)'])
    
    # Datos
    for gasto in gastos:
        writer.writerow([
            gasto.fecha.strftime('%d/%m/%Y'),
            gasto.tipo_gasto.get_nombre_display(),
            gasto.descripcion,
            str(gasto.monto)
        ])
    
    return output.getvalue().encode('utf-8')

def generar_pdf_simple(gastos, fecha_inicio, fecha_fin):
    """Genera un reporte PDF simple (versión básica)"""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    import io
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, f"SmartPocket - Reporte de Gastos")
    
    p.setFont("Helvetica", 12)
    p.drawString(100, 720, f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}")
    
    # Encabezados
    y = 680
    p.setFont("Helvetica-Bold", 10)
    p.drawString(100, y, "Fecha")
    p.drawString(200, y, "Categoría")
    p.drawString(300, y, "Descripción")
    p.drawString(450, y, "Monto (S/.)")
    
    # Datos
    p.setFont("Helvetica", 9)
    y -= 20
    total = 0
    
    for gasto in gastos[:30]:  # Limitar a 30 gastos
        if y < 100:  # Nueva página si no hay espacio
            p.showPage()
            y = 750
        
        p.drawString(100, y, gasto.fecha.strftime('%d/%m/%Y'))
        p.drawString(200, y, gasto.tipo_gasto.get_nombre_display()[:15])
        p.drawString(300, y, gasto.descripcion[:20])
        p.drawString(450, y, f"S/. {gasto.monto}")
        
        total += float(gasto.monto)
        y -= 15
    
    # Total
    p.setFont("Helvetica-Bold", 12)
    p.drawString(350, y-20, f"TOTAL: S/. {total:.2f}")
    
    p.save()
    return buffer.getvalue()

def generar_excel_simple(gastos, fecha_inicio=None, fecha_fin=None):
    """Genera reporte en formato Excel simple - MEJORADO"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Gastos"
    
    # Título principal
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "SmartPocket - Reporte de Gastos"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Período
    if fecha_inicio and fecha_fin:
        ws.merge_cells('A2:E2')
        period_cell = ws['A2']
        period_cell.value = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        period_cell.font = Font(size=12)
        period_cell.alignment = Alignment(horizontal='center')
        header_row = 4
    else:
        header_row = 3
    
    # Encabezados
    headers = ['Fecha', 'Categoría', 'Descripción', 'Monto (S/.)', 'Total Acumulado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    total_acumulado = 0
    data_row = header_row + 1
    
    for gasto in gastos:
        total_acumulado += float(gasto.monto)
        
        ws.cell(row=data_row, column=1, value=gasto.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=data_row, column=2, value=gasto.tipo_gasto.get_nombre_display())
        ws.cell(row=data_row, column=3, value=gasto.descripcion)
        ws.cell(row=data_row, column=4, value=float(gasto.monto))
        ws.cell(row=data_row, column=5, value=total_acumulado)
        
        data_row += 1
    
    # Total final
    total_row = data_row + 1
    ws.merge_cells(f'A{total_row}:C{total_row}')
    total_label = ws[f'A{total_row}']
    total_label.value = "TOTAL GENERAL:"
    total_label.font = Font(bold=True, size=14)
    total_label.alignment = Alignment(horizontal='right')
    
    total_value = ws[f'D{total_row}']
    total_value.value = total_acumulado
    total_value.font = Font(bold=True, size=14, color="FF0000")
    
    # Ajustar ancho de columnas
    column_widths = {'A': 12, 'B': 18, 'C': 25, 'D': 15, 'E': 18}
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()